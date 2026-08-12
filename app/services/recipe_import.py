"""
Bulk import of foundational batch/prep recipes from a restaurant's costing
workbook — the "GBB Batch Recipe Costing" format.

Written against a real workbook where every sheet (Meats, Sausages, Sides
and Carbs, Pickles, Rubs, Sauces and Salsas, Desserts) uses the same fixed
template: two recipe "cards" side by side (columns A-G and I-O), each
occupying an exact 27-row block regardless of how many ingredient rows are
actually used:

    row+0   title (col A/I), "Recipe Cost" label
    row+1   column headers: Ingredient | Measure | Procedure | RU | # of RU | RU Cost | Cost
    row+2..row+21   up to 20 ingredient rows
    row+22  "Tools/Equip:" ... "Total Cost:" ... <total cost>
    row+23  "Yield" <yield text> ... "# of RU:(<unit>)" ... <yield qty>
    row+24  "Shelf Life" <text> ... "RU Cost:" ... <cost per yield unit>
    row+25,26  blank gap before the next block

"RU" (recipe unit) is the costing unit for that ingredient line: almost
always "#" (kitchen shorthand for pound) or blank (also pound, by
convention in this workbook — verified against known per-pound prices),
occasionally a real unit like "g", "egg", "each". "# of RU" is the
quantity in that unit and "RU Cost" is the cost per unit — together
exactly what a RecipeIngredient needs (quantity + unit + implied cost).

These batch recipes routinely reference each other (a rub mix used across
several meats, a sauce used across several sides) under slightly different
names ("Brisket Rub Mix" as an ingredient vs. "Brisket Rub" as its own
card), so this module also proposes sub-recipe links using a conservative
normalized-exact match — never a fuzzy guess — leaving anything less
certain for a human to resolve in the review table before import.
"""

from __future__ import annotations
import io
import re
from typing import List, Dict, Optional
import openpyxl

BLOCK_HEIGHT = 27
CARD_COLUMN_OFFSETS = {"left": 0, "right": 8}  # column A=0 vs column I=8 (0-indexed)

# Ingredient-row unit shorthand seen in this workbook -> a real unit string.
_RU_UNIT_MAP = {
    "#": "lb", "lb": "lb", "g": "g", "egg": "egg", "head": "head",
    "pc": "each", "piece": "each", "slice": "slice", "qt": "qt",
}


class ImportParseError(Exception):
    pass


def _clean(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace(",", "").replace("$", "").strip()
        if not v:
            return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _norm_unit(ru: Optional[str]) -> str:
    if ru is None:
        return "lb"  # blank RU column defaults to pound in this workbook
    return _RU_UNIT_MAP.get(ru.strip().lower(), ru.strip().lower())


def _extract_yield_unit(ru_label: Optional[str]) -> Optional[str]:
    """'# of RU:(lb)' -> 'lb'; '# of RU: 1/3lb Links' -> '1/3lb Links'."""
    if not ru_label:
        return None
    s = re.sub(r'#\s*of\s*RU', '', ru_label, flags=re.IGNORECASE).replace(':', '').strip()
    m = re.search(r'\(([^)]+)\)', s)
    if m:
        s = m.group(1).strip()
    return s[:30] or None


def detect_workbook_type(file_bytes: bytes) -> str:
    """Returns 'batch_recipes' or 'menu_costing' based on which header row
    style the sheets use, or raises if neither matches."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)):
            vals = [_clean(c.value) for c in row]
            if vals and vals[0] == "Ingredient" and "Measure" in vals:
                return "batch_recipes"
            if vals and vals[0] == "Ingredient/Item" and "Amount" in vals:
                return "menu_costing"
    raise ImportParseError(
        "Couldn't recognize this workbook's layout. Expected either a batch recipe costing "
        "workbook ('Ingredient | Measure | Procedure | RU | # of RU | RU Cost | Cost' header) "
        "or a menu costing workbook ('Ingredient/Item | Amount | Cost' header)."
    )


def parse_batch_recipe_workbook(file_bytes: bytes) -> List[Dict]:
    """Parse every sheet into a flat list of recipe-card dicts:
    {recipe_name, sheet, yield_qty, yield_unit, yield_text, shelf_life,
     sheet_total_cost, sheet_ru_cost, ingredients: [
        {name, raw_measure, unit, quantity, unit_cost}, ...
     ]}
    Pure parsing only — no DB access, no sub-recipe/item resolution.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    recipes: List[Dict] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        max_row = ws.max_row
        grid: Dict[tuple, object] = {}
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
            for cell in row:
                grid[(cell.row, cell.column - 1)] = cell.value

        def g(r, c):
            return grid.get((r, c))

        block_start = 1
        while block_start <= max_row:
            for _card, coloff in CARD_COLUMN_OFFSETS.items():
                title = _clean(g(block_start, coloff))
                if not title or 'batch recipe cost card' in title.lower():
                    continue

                ingredients = []
                for r in range(block_start + 2, block_start + 22):
                    name = _clean(g(r, coloff))
                    if not name:
                        continue
                    qty = _num(g(r, coloff + 4))
                    unit_cost = _num(g(r, coloff + 5))
                    # Section-label rows ("Brine:", "Total Yield Weight:") carry
                    # no quantity/cost of their own — drop them rather than
                    # surface them as a fake zero-quantity ingredient.
                    if name.endswith(':') and qty is None and unit_cost is None:
                        continue
                    raw_measure = g(r, coloff + 1)
                    ru = _clean(g(r, coloff + 3))
                    ingredients.append({
                        "name": name,
                        "raw_measure": raw_measure if isinstance(raw_measure, (int, float)) else _clean(raw_measure),
                        "unit": _norm_unit(ru),
                        "quantity": qty,
                        "unit_cost": unit_cost,
                    })

                total_cost = _num(g(block_start + 22, coloff + 6))
                yield_text = _clean(g(block_start + 23, coloff + 1))
                ru_label = _clean(g(block_start + 23, coloff + 3))
                yield_qty = _num(g(block_start + 23, coloff + 6))
                shelf_life = _clean(g(block_start + 24, coloff + 1))
                ru_cost_final = _num(g(block_start + 24, coloff + 6))

                if not ingredients and total_cost is None and yield_qty is None:
                    continue  # empty placeholder card

                recipes.append({
                    "recipe_name": title,
                    "sheet": sheet_name,
                    "yield_qty": yield_qty,
                    "yield_unit": _extract_yield_unit(ru_label),
                    "yield_text": yield_text,
                    "shelf_life": shelf_life,
                    "sheet_total_cost": total_cost,
                    "sheet_ru_cost": ru_cost_final,
                    "ingredients": ingredients,
                })
            block_start += BLOCK_HEIGHT

    return recipes


def _norm_name(s: str) -> str:
    s = s.lower().strip()
    s = re.sub(r'\(.*?\)', '', s)          # drop parenthetical asides
    s = re.sub(r'[^a-z0-9 ]', ' ', s)      # drop punctuation
    s = re.sub(r'\bmix\b', ' ', s)         # "Brisket Rub Mix" ~ "Brisket Rub"
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def suggest_sub_recipe_links(recipes: List[Dict]) -> List[Dict]:
    """Mutates each ingredient dict in place, adding 'suggested_sub_recipe':
    the name of another card in this same workbook it normalized-exact
    matches, or None. Conservative on purpose — no fuzzy scoring — so a
    suggestion here is safe to default to without a human double-checking."""
    norm_titles: Dict[str, str] = {}
    for r in recipes:
        norm_titles.setdefault(_norm_name(r["recipe_name"]), r["recipe_name"])

    for r in recipes:
        self_norm = _norm_name(r["recipe_name"])
        for ing in r["ingredients"]:
            nn = _norm_name(ing["name"])
            match = norm_titles.get(nn)
            ing["suggested_sub_recipe"] = match if (match and nn != self_norm) else None

    return recipes


# ---------- Menu item costing workbook ----------
#
# A different template from the batch recipe workbook: each sheet repeats a
# variable-height card — title, "Ingredient/Item | Amount | Cost" header,
# ingredient rows (with blank-row gaps sprinkled throughout), then
# "Total Cost" / "Menu Price" / "Percentage Food Cost" immediately below
# the last ingredient (no fixed block height, unlike the batch workbook).
# Several cards sit side by side in independent column groups; parsed as a
# small state machine per group rather than assuming a row grid, which is
# what makes it robust to the ragged spacing in the real file. "Cost" here
# is each ingredient's *extended* cost (qty × unit cost), not a per-unit
# rate, so the per-unit cost is back-computed by dividing it out.

_MENU_FOOTER_LABELS = {"Total Cost", "Menu Price", "Percentage Food Cost"}


def _parse_amount(raw) -> tuple[Optional[float], Optional[str]]:
    """'1 lb'->(1.0,'lb'); '.5#'->(0.5,'lb'); 8.0 (bare number)->(8.0,'each');
    '' / None -> (None, None)."""
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return float(raw), "each"
    s = str(raw).strip()
    if not s:
        return None, None
    s = re.sub(r'\(.*?\)', '', s).strip()  # drop parenthetical asides
    m = re.match(r'^([\d.\/]+)\s*(.*)$', s)
    if not m:
        return None, (s or None)  # no leading number — keep the text as a hint
    num_str, unit_str = m.group(1), m.group(2).strip()
    try:
        if '/' in num_str:
            a, b = num_str.split('/')
            qty = float(a) / float(b)
        else:
            qty = float(num_str)
    except (ValueError, ZeroDivisionError):
        return None, (unit_str or s)
    unit = unit_str.rstrip('.').strip() or "each"
    if unit == '#':
        unit = 'lb'
    elif unit.lower() in ('pc', 'piece'):
        unit = 'each'
    return qty, unit


def _clean_category(sheet_name: str) -> str:
    s = re.sub(r'\s*X\s*$', '', sheet_name, flags=re.IGNORECASE).strip()
    return s or sheet_name.strip()


def parse_menu_costing_workbook(file_bytes: bytes) -> List[Dict]:
    """Parse every sheet (except one literally named "OLD", a deprecated
    sheet in the source workbook) into a flat list of menu-item dicts:
    {recipe_name, sheet, category, menu_price, sheet_total_cost,
     ingredients: [{name, raw_amount, quantity, unit, unit_cost}, ...]}.
    Pure parsing only — no DB access, no item/sub-recipe resolution.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    recipes: List[Dict] = []

    for sheet_name in wb.sheetnames:
        if sheet_name.strip().lower() == "old":
            continue
        ws = wb[sheet_name]
        max_row = ws.max_row
        grid: Dict[tuple, object] = {}
        for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=False):
            for cell in row:
                grid[(cell.row, cell.column - 1)] = cell.value

        def g(r, c):
            return grid.get((r, c))

        header_cols = sorted({c for (r, c), v in grid.items() if _clean(v) == "Ingredient/Item"})
        if not header_cols:
            continue
        category = _clean_category(sheet_name)

        for offset in header_cols:
            r = 1
            while r <= max_row:
                raw0 = g(r, offset)
                name0 = _clean(raw0)
                if not name0:
                    r += 1
                    continue
                if not isinstance(raw0, str):
                    r += 1
                    continue  # stray numeric annotation in this column, not a title
                if name0 == "Ingredient/Item" or name0 in _MENU_FOOTER_LABELS:
                    r += 1
                    continue

                title = name0  # candidate card title
                r += 1
                if _clean(g(r, offset)) == "Ingredient/Item":
                    r += 1  # skip the repeated column header

                ingredients = []
                total_cost = None
                while r <= max_row:
                    raw_cell0 = g(r, offset)
                    cell0 = _clean(raw_cell0)
                    if cell0 is None:
                        r += 1
                        continue
                    if not isinstance(raw_cell0, str):
                        r += 1
                        continue
                    if cell0 == "Total Cost":
                        total_cost = _num(g(r, offset + 2))
                        r += 1
                        break
                    raw_amt = g(r, offset + 1)
                    ext_cost = _num(g(r, offset + 2))
                    qty, unit = _parse_amount(raw_amt)
                    if qty and qty > 0:
                        unit_cost = round(ext_cost / qty, 6) if ext_cost is not None else None
                    else:
                        # no parseable quantity ("Pickle Avg" style flat-rate line) —
                        # treat as one flat-cost unit rather than dropping it.
                        qty = 1.0
                        unit = unit or "each"
                        unit_cost = ext_cost
                    ingredients.append({
                        "name": cell0, "raw_amount": raw_amt,
                        "quantity": qty, "unit": unit, "unit_cost": unit_cost,
                    })
                    r += 1

                menu_price = None
                if r <= max_row and _clean(g(r, offset)) == "Menu Price":
                    menu_price = _num(g(r, offset + 2))
                    r += 1
                if r <= max_row and _clean(g(r, offset)) == "Percentage Food Cost":
                    r += 1  # informational only — the app recomputes this live

                if not ingredients and total_cost is None and menu_price is None:
                    continue  # nothing usable in this card

                recipes.append({
                    "recipe_name": title,
                    "sheet": sheet_name,
                    "category": category,
                    "menu_price": menu_price,
                    "sheet_total_cost": total_cost,
                    "ingredients": ingredients,
                })

    return recipes
